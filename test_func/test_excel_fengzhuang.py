# 封装一个可以任务读取数据的py文件
import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        row = list(sheet.rows)
        title = [i.value for i in row[0]]
        cases = []
        # 获取除了第一行之外的数据
        for item in row[1:]:
            # 获取对应的值
            data = [i.value for i in item]
            # 字典打包
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases


    def data_write(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
